from django.contrib.auth import logout
from django.urls import reverse
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from backend.models import Asset,AssetCategory,Report,InactiveAsset
from django.core.mail import send_mail
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib import messages
import pandas as pd
from backend.forms import AssetUploadForm
from datetime import datetime
from openpyxl import load_workbook
import os
from django.http import HttpResponse
from django.conf import settings
from django.http import FileResponse
from openpyxl import Workbook
import pytz
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import re
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


# Create your views here.
@login_required
def dashboard(request):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    assets = Asset.objects.order_by('-item_creation_date')[:5]
    a = Asset.objects.filter(is_active=True)
    # assets = Asset.objects.filter(is_active=True).order_by('-item_creation_date')[:5]
    category = AssetCategory.objects.all()
    total_category = category.count()
    total_assets = a.count()
    total_owners = Asset.objects.values('owner').distinct().count()
    context = {
        'assets': assets,
        'total_assets': total_assets,
        'total_category': total_category,
        'category': category,
        'a': a,
        'total_owners': total_owners,
        'role_name': role_name,
     }
    return render(request, 'index.html',context)

from django.db.models import Q

# manage asset view
@login_required
def manageasset(request):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    search_term = request.GET.get('search_term')
    # assets = Asset.objects.order_by('-item_creation_date')
    assets = Asset.objects.filter(is_active=True).order_by('-item_creation_date')


    if search_term:
        assets = assets.filter(Q(asset_id__icontains=search_term))
        

    paginator = Paginator(assets, 10)  # Show 5 assets per page

    page = request.GET.get('page')
    try:
        assets = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        assets = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        assets = paginator.page(paginator.num_pages)

    # Calculate the range of pages to display
    num_pages = paginator.num_pages
    current_page = assets.number

    if num_pages <= 5:
        page_range = range(1, num_pages + 1)
    elif current_page <= 3:
        page_range = range(1, 6)
    elif current_page >= num_pages - 2:
        page_range = range(num_pages - 4, num_pages + 1)
    else:
        page_range = range(current_page - 2, current_page + 3)

    context = {
        'assets': assets,
        'category': AssetCategory.objects.all(),
        'page_range': page_range,
        'search_term': search_term,
        'role_name' : role_name,
    }

    return render(request, 'tables.html', context)


# category view with pagination
@login_required
def category(request):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    category = AssetCategory.objects.all()
    categories_list = AssetCategory.objects.order_by('-item_creation_date')
    paginator = Paginator(categories_list, 10) # Show 10 categories per page

    page = request.GET.get('page')
    try:
        categories = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        categories = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        categories = paginator.page(paginator.num_pages)

    context = {
        'categories': categories,
        'category':category,
        'role_name' : role_name,
    }
    return render(request, 'category.html', context)


# add category
@login_required
def add_category(request):
    if request.method == 'POST':
        category_name = request.POST.get('category_name')

        # Check if the category name already exists
        if AssetCategory.objects.filter(category_name=category_name).exists():
            messages.error(request, 'Category already exists')
        else:
            category = AssetCategory.objects.create(category_name=category_name)
            messages.success(request, 'Category added successfully')

        return redirect('backend:category')
    else:
        return render(request, 'category.html')

#delete category
@login_required
def delete_category(request, id):
    c = get_object_or_404(AssetCategory, id=id)
    c.delete()

    messages.success(request, 'Category deleted successfully')
    return redirect(reverse('backend:category'))

# edit category
@login_required
def edit_category(request):
    if request.method == 'POST':
        c_id = request.POST['id']
        cat = get_object_or_404(AssetCategory, id=c_id)
        cat.category_name = request.POST['category_name']
        cat.save()

        messages.success(request, 'Category updated successfully')
        return redirect('backend:category')

    return render(request, 'category.html')

# same category asset table view
@login_required
def category_table(request, category_id):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    category = AssetCategory.objects.get(id=category_id)
    asset_list = Asset.objects.filter(category=category, is_active=True)  # Filter for active assets

    # search functionality
    search_term = request.GET.get('search_term')
    if search_term:
        asset_list = asset_list.filter(asset_id__icontains=search_term.lower())

    paginator = Paginator(asset_list, 10)  # Show 10 assets per page
    page = request.GET.get('page')
    assets = paginator.get_page(page)

    # Calculate the range of pages to display
    num_pages = paginator.num_pages
    current_page = assets.number

    if num_pages <= 5:
        page_range = range(1, num_pages + 1)
    elif current_page <= 3:
        page_range = range(1, 6)
    elif current_page >= num_pages - 2:
        page_range = range(num_pages - 4, num_pages + 1)
    else:
        page_range = range(current_page - 2, current_page + 3)

    category_name = category.category_name
    category = AssetCategory.objects.all()

    context = {
        'category': category,
        'assets': assets,
        'category_name': category_name,
        'page_range': page_range,
        'role_name': role_name,
    }

    return render(request, 'category_table.html', context)

# @login_required
# def category_table(request, category_id):
#     category = AssetCategory.objects.get(id=category_id)
#     asset_list = Asset.objects.filter(category=category)

#     # search functionality
#     search_term = request.GET.get('search_term')
#     if search_term:
#         asset_list = asset_list.filter(asset_id__icontains=search_term.lower())
        

#     paginator = Paginator(asset_list, 10)  # Show 10 assets per page
#     page = request.GET.get('page')
#     assets = paginator.get_page(page)

#     # Calculate the range of pages to display
#     num_pages = paginator.num_pages
#     current_page = assets.number

#     if num_pages <= 5:
#         page_range = range(1, num_pages + 1)
#     elif current_page <= 3:
#         page_range = range(1, 6)
#     elif current_page >= num_pages - 2:
#         page_range = range(num_pages - 4, num_pages + 1)
#     else:
#         page_range = range(current_page - 2, current_page + 3)

#     category_name = category.category_name
#     category = AssetCategory.objects.all()

#     context = {
#         'category': category,
#         'assets': assets,
#         'category_name': category_name,
#         'page_range': page_range,
#     }
    
#     return render(request, 'category_table.html', context)



# report view
@login_required
def report(request):
    category = AssetCategory.objects.all()
    user = request.user

    show_reports = False
    hide_action_column = False  # Flag to determine if the action column should be hidden
    role_name = None

    if user.is_superuser and user.is_staff:
        # Show all reports for superadmin and staff users
        report_list = Report.objects.order_by('item_creation_date').all()
        show_reports = True
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        # Show all reports for staff users who are not in any group
        report_list = Report.objects.order_by('item_creation_date').all()
        show_reports = True
        hide_action_column = True  # Hide the action column for these users
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        # Show only estate service reports for staff users in state_user group
        report_list = Report.objects.filter(service_type='estate').order_by('item_creation_date').all()
        show_reports = True
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        # Show only ICT service reports for staff users in ict_user group
        report_list = Report.objects.filter(service_type='ict').order_by('item_creation_date').all()
        show_reports = True
        role_name = 'ICT'
    else:
        # For other users, show no reports
        report_list = Report.objects.none()

    # Handle search functionality
    search_term = request.GET.get('search_term')
    if search_term:
        report_list = report_list.filter(asset__asset_id__icontains=search_term)

    paginator = Paginator(report_list, 10)  # Show 10 reports per page
    page = request.GET.get('page')
    reports = paginator.get_page(page)

    # Calculate the range of pages to display
    num_pages = paginator.num_pages
    current_page = reports.number

    if num_pages <= 5:
        page_range = range(1, num_pages + 1)
    elif current_page <= 3:
        page_range = range(1, 6)
    elif current_page >= num_pages - 2:
        page_range = range(num_pages - 4, num_pages + 1)
    else:
        page_range = range(current_page - 2, current_page + 3)

    prev_page = reports.previous_page_number if reports.has_previous() else None
    next_page = reports.next_page_number if reports.has_next() else None

    context = {
        'category': category,
        'reports': reports,
        'search_term': search_term,
        'page_range': page_range,
        'prev_page': prev_page,
        'next_page': next_page,
        'show_reports': show_reports,
        'hide_action_column': hide_action_column,  # Pass the flag to the template
        'user': user,
        'role_name': role_name,  # Pass the role name to the template
    }

    return render(request, 'report.html', context)

# def report(request):
#     category = AssetCategory.objects.all()
#     user = request.user

#     show_reports = False
#     hide_action_column = False  # Flag to determine if the action column should be hidden

#     if user.is_superuser and user.is_staff:
#         # Show all reports for superadmin and staff users
#         report_list = Report.objects.order_by('item_creation_date').all()
#         show_reports = True
#     elif user.is_staff and not user.groups.exists():
#         # Show all reports for staff users who are not in any group
#         report_list = Report.objects.order_by('item_creation_date').all()
#         show_reports = True
#         hide_action_column = True  # Hide the action column for these users
#     elif user.groups.filter(name='state_user').exists():
#         # Show only estate service reports for staff users in state_user group
#         report_list = Report.objects.filter(service_type='estate').order_by('item_creation_date').all()
#         show_reports = True
#     elif user.groups.filter(name='ict_user').exists():
#         # Show only ICT service reports for staff users in ict_user group
#         report_list = Report.objects.filter(service_type='ict').order_by('item_creation_date').all()
#         show_reports = True
#     else:
#         # For other users, show no reports
#         report_list = Report.objects.none()

#     # Handle search functionality
#     search_term = request.GET.get('search_term')
#     if search_term:
#         report_list = report_list.filter(asset__asset_id__icontains=search_term)

#     paginator = Paginator(report_list, 10)  # Show 10 reports per page
#     page = request.GET.get('page')
#     reports = paginator.get_page(page)

#     # Calculate the range of pages to display
#     num_pages = paginator.num_pages
#     current_page = reports.number

#     if num_pages <= 5:
#         page_range = range(1, num_pages + 1)
#     elif current_page <= 3:
#         page_range = range(1, 6)
#     elif current_page >= num_pages - 2:
#         page_range = range(num_pages - 4, num_pages + 1)
#     else:
#         page_range = range(current_page - 2, current_page + 3)

#     prev_page = reports.previous_page_number if reports.has_previous() else None
#     next_page = reports.next_page_number if reports.has_next() else None

#     context = {
#         'category': category,
#         'reports': reports,
#         'search_term': search_term,
#         'page_range': page_range,
#         'prev_page': prev_page,
#         'next_page': next_page,
#         'show_reports': show_reports,
#         'hide_action_column': hide_action_column,  # Pass the flag to the template
#         'user' : user, 
#     }

#     return render(request, 'report.html', context)

# @login_required
# def report(request):
#     category = AssetCategory.objects.all()
#     user = request.user
    
#     show_reports = False

#     if user.is_superuser and user.is_staff:
#         # Show all reports for superadmin and staff users
#         report_list = Report.objects.order_by('item_creation_date').order_by('item_creation_date').all()
#         show_reports = True
#     elif user.is_staff and not user.groups.exists():
#     # Show all reports for superadmin and staff users
#         report_list = Report.objects.order_by('item_creation_date').order_by('item_creation_date').all()
#         show_reports = True

#     elif user.groups.filter(name='state_user').exists():
#         # Show only estate service reports for staff users in state_user group
#         report_list = Report.objects.filter(service_type='estate').order_by('item_creation_date').all()
#         show_reports = True
#     elif user.groups.filter(name='ict_user').exists():
#         # Show only ICT service reports for staff users in ict_user group
#         report_list = Report.objects.filter(service_type='ict').order_by('item_creation_date').all()
#         show_reports = True
#     else:
#         # For other users, show no reports
#         report_list = Report.objects.none()

#     # Handle search functionality
#     search_term = request.GET.get('search_term')
#     if search_term:
#         report_list = report_list.filter(asset__asset_id__icontains=search_term)

#     paginator = Paginator(report_list, 10)  # Show 10 reports per page
#     page = request.GET.get('page')
#     reports = paginator.get_page(page)

#     # Calculate the range of pages to display
#     num_pages = paginator.num_pages
#     current_page = reports.number

#     if num_pages <= 5:
#         page_range = range(1, num_pages + 1)
#     elif current_page <= 3:
#         page_range = range(1, 6)
#     elif current_page >= num_pages - 2:
#         page_range = range(num_pages - 4, num_pages + 1)
#     else:
#         page_range = range(current_page - 2, current_page + 3)

#     prev_page = reports.previous_page_number if reports.has_previous() else None
#     next_page = reports.next_page_number if reports.has_next() else None

#     context = {
#         'category': category,
#         'reports': reports,
#         'search_term': search_term,
#         'page_range': page_range,
#         'prev_page': prev_page,
#         'next_page': next_page,
#         'show_reports': show_reports,
#     }

#     return render(request, 'report.html', context)




# report remark view
@login_required
def report_remark(request):
    if request.method == 'POST':
        asset_id = request.POST['id']
        r = get_object_or_404(Report, asset_id=asset_id)
        r.remark = request.POST['remark']
        r.save()
        
        messages.success(request, 'Remark updated successfully')
        return redirect('backend:report')


# delete report 
@login_required
def delete_report(request, report_id):
    if request.method == 'POST':
        report = get_object_or_404(Report, id=report_id)
        report.delete()

        messages.success(request, 'Report deleted successfully')
        return redirect('backend:report')
    

# logout view
@login_required
def logout_view(request):
    logout(request)
    return redirect('frontend:admin_login')



# password change
@login_required
def profile(request):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid() and (user.is_staff or user.is_superuser):
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect(reverse('backend:profile'))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    if user.is_staff or user.is_superuser:
        category = AssetCategory.objects.all()
        context = {
            'form': form,
            'category': category,
            'role_name': role_name,
            'user' : user,
        }
        return render(request, 'profile.html', context)
    else:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect(reverse('backend:profile'))
# def profile(request):
#     user = request.user
#     if request.method == 'POST':
#         form = PasswordChangeForm(request.user, request.POST)
#         if form.is_valid() and (request.user.is_staff or request.user.is_superuser):
#             user = form.save()
#             update_session_auth_hash(request, user)  # Important!
#             messages.success(request, 'Your password was successfully updated!')
#             return redirect(reverse('backend:profile'))
#         else:
#             for field, errors in form.errors.items():
#                 for error in errors:
#                     messages.error(request, f"{field.capitalize()}: {error}")
#             messages.error(request, 'Please correct the errors below.')
#     else:
#         form = PasswordChangeForm(request.user)
    
#     if request.user.is_staff or request.user.is_superuser:
#         category = AssetCategory.objects.all()
#         return render(request, 'profile.html', {'form': form, 'category': category})
#     else:
#         messages.error(request, 'You do not have permission to access this page.')
#         return redirect(reverse('backend:profile'))

# username and email change


@login_required
def update_admin_profile(request):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    if request.method == 'POST':
        if user.is_staff or user.is_superuser:
            old_username = user.username
            old_email = user.email

            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.save()

            if user.username != old_username:
                messages.success(request, 'Username has been changed successfully.')
            if user.email != old_email:
                messages.success(request, 'Email has been changed successfully.')

            return redirect(reverse('backend:profile'))
    
    context = {
        'user': user,
        'role_name': role_name,
    }
    return render(request, 'profile.html', context)
# def update_admin_profile(request):
#     if request.method == 'POST':
#         user = request.user
#         if user.is_staff or user.is_superuser:
#             old_username = user.username
#             old_email = user.email

#             user.username = request.POST.get('username')
#             user.email = request.POST.get('email')
#             user.save()

#             if user.username != old_username:
#                 messages.success(request, 'Username has been changed successfully.')
#             if user.email != old_email:
#                 messages.success(request, 'Email has been changed successfully.')

#             return redirect(reverse('backend:profile'))
    
#     return render(request, 'profile.html')



# list the users baased table based on role
@login_required
def user_list(request):
    if request.user.is_superuser:
        query = request.GET.get('q')  # Get the search query from the request
        user_list = User.objects.all().order_by('-id')

        if query:
            user_list = user_list.filter(Q(username__icontains=query))  # Filter users by username

        paginator = Paginator(user_list, 10)  # Show 10 users per page

        page = request.GET.get('page')
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            users = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            users = paginator.page(paginator.num_pages)

        category = AssetCategory.objects.all()
        context = {'users': users, 'category': category, 'search_query': query}
        return render(request, 'users.html', context)
    else:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect(reverse('backend:dashboard'))

# delete users
@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    return redirect(reverse('backend:user_list'))

# add new users based on roles
@login_required
def add_user(request):
    if request.method == 'POST':
        # get form data
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        role = request.POST.get('role')

        # validate form data
        if password != confirm_password:
            messages.error(request, 'Passwords do not match')
            return redirect('add_user')

        # create new user
        new_user = User.objects.create_user(username=username, email=email, password=password)

        # add role to new user
        if role == 'admin':
            new_user.is_superuser = True
            new_user.is_staff = True
            new_user.save()
        elif role == 'staff_user':
            new_user.is_staff = True
            new_user.save()
        elif role == 'state_user':
            # Create the "state_user" group if it doesn't exist
            group, created = Group.objects.get_or_create(name='state_user')
            new_user.groups.add(group)
            new_user.is_staff = True
            new_user.save()
        elif role == 'ICT_user':
            # Create the "ict_user" group if it doesn't exist
            group, created = Group.objects.get_or_create(name='ict_user')
            new_user.groups.add(group)
            new_user.is_staff = True
            new_user.save()
        else:
            messages.error(request, 'Invalid role')
            return redirect(reverse('backend:user_list'))

        messages.success(request, 'User created successfully')
        return redirect(reverse('backend:user_list'))

    return render(request, 'users.html', {'user_created': False})


# adding asset details individually
@login_required
def add_asset(request):
    if request.method == 'POST':
        asset_id = request.POST['asset_id']
        name = request.POST['name']
        category_id = request.POST['category']
        category = AssetCategory.objects.get(id=category_id)
        # sub_category = request.POST['sub_category']
        location = request.POST['location']
        owner = request.POST['owner']
        purchase_date = request.POST['purchase_date']
        
        # Check if asset_id already exists
        if Asset.objects.filter(asset_id=asset_id).exists():
            messages.error(request, 'Asset with the same ID already exists.')
            return redirect(reverse('backend:manageasset'))
        
        asset = Asset(
            asset_id=asset_id,
            name=name,
            category=category,
            # sub_category=sub_category,
            location=location,
            owner=owner,
            purchase_date=purchase_date
        )
        asset.save()
        messages.success(request, 'Asset added successfully.')
        return redirect(reverse('backend:manageasset'))
    
    return render(request, 'tables.html')




# edit asset details
@login_required
def edit_asset(request):
    if request.method == 'POST':
        asset_id = request.POST['asset_id']
        asset = Asset.objects.get(asset_id=asset_id)
        asset.name = request.POST['name']
        category_id = request.POST['category']
        asset.category = AssetCategory.objects.get(id=category_id)
        # asset.sub_category = request.POST['sub_category']
        asset.location = request.POST['location']
        asset.owner = request.POST['owner']
        asset.purchase_date = request.POST['purchase_date']
        asset.save()

        messages.success(request, 'Asset updated successfully')

        source_page = request.POST['source_page']
        if source_page == 'inactive':
            return redirect(reverse('backend:inactive_assets'))
        elif source_page == 'active':
            return redirect(reverse('backend:manageasset'))
        else:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        

    return render(request, 'tables.html')

# @login_required
# def edit_asset(request):
#     if request.method == 'POST':
#         asset_id = request.POST['asset_id']
#         asset = Asset.objects.get(asset_id=asset_id)
#         asset.name = request.POST['name']
#         category_id = request.POST['category']
#         asset.category = AssetCategory.objects.get(id=category_id)
#         # asset.sub_category = request.POST['sub_category']
#         asset.location = request.POST['location']
#         asset.owner = request.POST['owner']
#         asset.purchase_date = request.POST['purchase_date']
#         asset.save()
        
#         messages.success(request, 'Asset updated successfully')
#         return redirect(reverse('backend:manageasset'))
    
#     return render(request, 'tables.html')


# deactivate
@login_required
def transfer_asset(request, asset_id):
    asset = get_object_or_404(Asset, asset_id=asset_id)
    inactive_asset = InactiveAsset(asset=asset)
    inactive_asset.save()
    asset.is_active = False
    asset.save()

    messages.success(request, 'Asset has been removed from active listing.')
    # Redirect back to the current page
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
# @login_required
# def transfer_asset(request, asset_id):
#     asset = get_object_or_404(Asset, asset_id=asset_id)
#     inactive_asset = InactiveAsset(asset=asset)
#     inactive_asset.save()
#     asset.is_active = False
#     asset.save()

    # messages.success(request, 'Asset has been removed from active listing.')
    # return redirect('backend:manageasset')


from django.http import HttpResponseRedirect
# delete asset
@login_required 
def delete_asset(request, asset_id):
    asset = get_object_or_404(Asset, asset_id=asset_id)
    asset.delete()
    
    messages.success(request, 'Asset deleted successfully')
    
    # Get the 'next' parameter from the request's POST data
    next_url = request.POST.get('next')
    
    # Redirect the user to the 'next' URL if it exists, otherwise redirect to a default URL
    if next_url:
        return HttpResponseRedirect(next_url)
    else:
        return redirect(reverse('backend:manageasset'))
# @login_required 
# def delete_asset(request, asset_id):
#     asset = get_object_or_404(Asset, asset_id=asset_id)
#     asset.delete()
    
#     messages.success(request, 'Asset deleted successfully')
#     return redirect(reverse('backend:manageasset'))


# view of report of faulty asset
def save_report(request):
    if request.method == 'POST':
        # Get the form data
        asset_id = request.POST['asset_id']
        asset = Asset.objects.get(asset_id=asset_id)
        name = request.POST['name']
        email = request.POST['email']
        service_type = request.POST['service']  # Get the selected service type
        description = request.POST['description']

        # Check if a report object already exists for this asset
        if Report.objects.filter(asset=asset).exists():
            return redirect(reverse('frontend:report_error'))

        # Create a new report object
        report = Report.objects.create(
            name=name,
            email=email,
            service_type=service_type,  # Save the selected service type
            description=description,
            asset=asset
        )

        # Get the group name based on the service type
        group_name = 'state_user' if service_type == 'estate' else 'ict_user'

        # Get the users in the group
        group = Group.objects.get(name=group_name)
        users = group.user_set.all()

        # Retrieve the admin user
        admin_user = User.objects.filter(is_superuser=True).first()

        # Exclude the admin user from the recipients
        users = users.exclude(id=admin_user.id)

        # Send email to the users in the group
        user_emails = list(users.values_list('email', flat=True))
        send_mail(
            'New Report Submitted',
            f'A new report has been submitted for asset id {asset_id}.\n\nName: {name}\nEmail: {email}\nService Type: {service_type}\nDescription: {description}',
            'from@example.com',
            user_emails,
            fail_silently=False,
        )

        return redirect('frontend:index')

    return render(request, 'frontend/detail.html')

import tempfile
import re

@login_required
def upload_assets(request):
    form = AssetUploadForm()

    if request.method == 'POST':
        form = AssetUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']

            # Save the uploaded file to a temporary location
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(file.read())

            df = pd.read_excel(temp_file.name)

            existing_asset_ids = Asset.objects.values_list('asset_id', flat=True)

            duplicate_assets = []
            new_assets = []

            validation_errors = []
            invalid_asset_ids = []  # Store invalid asset_id values
            invalid_name_rows = []  # Store rows with invalid name format
            invalid_location_rows = []  # Store rows with invalid location format
            invalid_owner_rows = []  # Store rows with invalid owner format

            for _, row in df.iterrows():
                asset_id = str(row['asset_id'])  # Convert to string
                if asset_id in existing_asset_ids:
                    duplicate_assets.append(asset_id)
                    continue

                asset_category, _ = AssetCategory.objects.get_or_create(category_name=row['category'])

                # Validate asset_id using regex
                if not re.match(r'^GC\d{2}\/\d{2}\/\d{5}$', str(asset_id)):
                    invalid_asset_ids.append(str(asset_id))  # Add invalid asset_id to the list
                    continue

                # Validate name using regex
                name = row['name']
                if not re.match(r'^[A-Za-z\s]+$', name):
                    invalid_name_rows.append(row)  # Add the row to invalid_name_rows
                    continue

                # Validate location using regex
                location = row['location']
                if not re.match(r'^[A-Za-z0-9\s]+$', str(location)):
                    invalid_location_rows.append(row)  # Add the row to invalid_location_rows
                    continue

                # Validate owner using regex
                owner = row['owner']
                if not re.match(r'^(RUB\d{9}|GCIT\d{8}|\d{8}|\d{9})$', str(owner)):
                    invalid_owner_rows.append(row)  # Add the row to invalid_owner_rows
                    continue

                purchase_date_str = str(row['purchase_date'])
                purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                asset = Asset(
                    asset_id=asset_id,
                    name=name,
                    category=asset_category,
                    location=location,
                    owner=owner,
                    purchase_date=purchase_date
                )
                new_assets.append(asset)

            if duplicate_assets:
                messages.warning(request, "The file contains assets that already exist in the system")
                return redirect(reverse('backend:manageasset'))

            if invalid_asset_ids:
                invalid_asset_ids_msg = ', '.join(invalid_asset_ids)  # Join invalid asset_id values
                messages.error(request, f"Invalid asset_id format for assets: {invalid_asset_ids_msg}")
                return redirect(reverse('backend:manageasset'))

            if invalid_name_rows:
                validation_errors.append("Invalid name format for the following assets:")
                for row in invalid_name_rows:
                    asset_id = row['asset_id']
                    validation_errors.append(f"- Asset ID: {asset_id}")
                validation_errors.append("")  # Add a blank line for formatting
            
            if invalid_location_rows:
                validation_errors.append("Invalid location format for the following assets:")
                for row in invalid_location_rows:
                    asset_id = row['asset_id']
                    validation_errors.append(f"- Asset ID: {asset_id}")
                validation_errors.append("")  # Add a blank line for formatting

            if invalid_owner_rows:
                validation_errors.append("Invalid owner format for the following assets:")
                for row in invalid_owner_rows:
                    asset_id = row['asset_id']
                    validation_errors.append(f"- Asset ID: {asset_id}")
                validation_errors.append("")  # Add a blank line for formatting

            if validation_errors:
                for error in validation_errors:
                    messages.error(request, error)
                return redirect(reverse('backend:manageasset'))

            if new_assets:
                Asset.objects.bulk_create(new_assets)

                # Generate QR code and save for each asset
                for asset in new_assets:
                    asset.save()

                messages.success(request, 'File Uploaded Successfully')

            # Delete the temporary file
            os.remove(temp_file.name)

            return redirect(reverse('backend:manageasset'))

    return render(request, 'tables.html', {'form': form})



# def upload_assets(request):
#     form = AssetUploadForm()

#     if request.method == 'POST':
#         form = AssetUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = form.cleaned_data['file']

#             # Save the uploaded file to a temporary location
#             with tempfile.NamedTemporaryFile(delete=False) as temp_file:
#                 temp_file.write(file.read())

#             df = pd.read_excel(temp_file.name)

#             existing_asset_ids = Asset.objects.values_list('asset_id', flat=True)

#             duplicate_assets = []
#             new_assets = []

#             validation_errors = []
#             invalid_asset_ids = []  # Store invalid asset_id values

#             for _, row in df.iterrows():
#                 asset_id = str(row['asset_id'])  # Convert to string
#                 if asset_id in existing_asset_ids:
#                     duplicate_assets.append(asset_id)
#                     continue

#                 asset_category, _ = AssetCategory.objects.get_or_create(category_name=row['category'])

#                 # Validate asset_id using regex
#                 if not re.match(r'^GC\d{2}\/\d{2}\/\d{5}$', str(asset_id)):
#                     invalid_asset_ids.append(str(asset_id))  # Add invalid asset_id to the list
#                     continue

#                 # Validate name using regex
#                 name = row['name']
#                 if not re.match(r'^[A-Za-z\s]+$', name):
#                     validation_errors.append(f"Invalid name format for asset: {asset_id}")
#                     break  # Stop processing further rows

#                 # Validate location using regex
#                 location = row['location']
#                 if not re.match(r'^[A-Za-z0-9\s]+$', str(location)):
#                     validation_errors.append(f"Invalid location format for asset: {asset_id}")
#                     break  # Stop processing further rows

#                 # Validate owner using regex
#                 owner = row['owner']
#                 if not re.match(r'^(RUB\d{9}|GCIT\d{8}|\d{8}|\d{9})$', str(owner)):
#                     validation_errors.append(f"Invalid owner format for asset: {asset_id}")
#                     break  # Stop processing further rows
                

#                 purchase_date_str = str(row['purchase_date'])
#                 purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
#                 asset = Asset(
#                     asset_id=asset_id,
#                     name=name,
#                     category=asset_category,
#                     location=location,
#                     owner=owner,
#                     purchase_date=purchase_date
#                 )
#                 new_assets.append(asset)

#             if duplicate_assets:
#                 messages.warning(request, "The file contains assets that already exist in the system")
#                 return redirect(reverse('backend:manageasset'))

#             if invalid_asset_ids:
#                 invalid_asset_ids_msg = ', '.join(invalid_asset_ids)  # Join invalid asset_id values
#                 validation_errors.append(f"Invalid asset_id format for assets: {invalid_asset_ids_msg}")
#                 for error in validation_errors:
#                     messages.error(request, error)
#                 return redirect(reverse('backend:manageasset'))

#             if new_assets:
#                 Asset.objects.bulk_create(new_assets)

#                 # Generate QR code and save for each asset
#                 for asset in new_assets:
#                     asset.save()

#                 messages.success(request, 'File Uploaded Successfully')

#             # Delete the temporary file
#             os.remove(temp_file.name)

#             return redirect(reverse('backend:manageasset'))

#     return render(request, 'tables.html', {'form': form})




# def upload_assets(request):
#     form = AssetUploadForm()

#     if request.method == 'POST':
#         form = AssetUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = form.cleaned_data['file']
#             df = pd.read_excel(file)

#             existing_asset_ids = Asset.objects.values_list('asset_id', flat=True)

#             duplicate_assets = []
#             new_assets = []

#             for _, row in df.iterrows():
#                 asset_id = row['asset_id']
#                 if asset_id in existing_asset_ids:
#                     duplicate_assets.append(asset_id)
#                 else:
#                     asset_category, _ = AssetCategory.objects.get_or_create(category_name=row['category'])
#                     purchase_date_str = str(row['purchase_date'])
#                     purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
#                     asset = Asset(
#                         asset_id=asset_id,
#                         name=row['name'],
#                         category=asset_category,
#                         location=row['location'],
#                         owner=row['owner'],
#                         purchase_date=purchase_date
#                     )
#                     new_assets.append(asset)

#             if duplicate_assets:
#                 messages.warning(request, "The file contains asset that already exist in the system")
#                 return redirect(reverse('backend:manageasset'))
            
#             if new_assets:
#                 Asset.objects.bulk_create(new_assets)

#                 # Generate QR code and save for each asset
#                 for asset in new_assets:
#                     asset.save()

#                 messages.success(request, 'File Uploaded Successfully')

#             return redirect(reverse('backend:manageasset'))

#     return render(request, 'tables.html', {'form': form})

# @login_required
# def upload_assets(request):
#     form = AssetUploadForm()

#     if request.method == 'POST':
#         form = AssetUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = form.cleaned_data['file']
#             df = pd.read_excel(file)

#             existing_asset_ids = Asset.objects.values_list('asset_id', flat=True)

#             duplicate_assets = []
#             new_assets = []

#             for _, row in df.iterrows():
#                 asset_id = row['asset_id']
#                 if asset_id in existing_asset_ids:
#                     duplicate_assets.append(asset_id)
#                 else:
#                     asset_category, _ = AssetCategory.objects.get_or_create(category_name=row['category'])
#                     purchase_date_str = str(row['purchase_date'])
#                     purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
#                     asset = Asset(
#                         asset_id=asset_id,
#                         name=row['name'],
#                         category=asset_category,
#                         location=row['location'],
#                         owner=row['owner'],
#                         purchase_date=purchase_date
#                     )
#                     new_assets.append(asset)

#             if duplicate_assets:
#                 messages.warning(request, "The file contains some duplicate assets")

#             if new_assets:
#                 Asset.objects.bulk_create(new_assets)

#                 # Generate QR code and save for each asset
#                 for asset in new_assets:
#                     asset.save()

#                 messages.success(request, 'File Uploaded Successfully')

#             return redirect(reverse('backend:manageasset'))  

#     return render(request, 'tables.html', {'form': form})




# export asset details in excel format
@login_required
def export_to_excel(request):
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Write the headers for the Excel file
    headers = ['asset_id', 'name', 'category', 'location', 'owner', 'purchase_date']
    ws.append(headers)
    
    # Get the active assets
    assets = Asset.objects.filter(is_active=True).values_list(
        'asset_id', 'name', 'category__category_name', 'location', 'owner', 'purchase_date'
    )
    
    # Write the asset data to the Excel file
    for asset in assets:
        ws.append(asset)
    
    # Save the Excel file to a specific location
    file_path = os.path.join(settings.BASE_DIR, 'assets.xlsx')
    wb.save(file_path)
    
    # Check if the file exists
    if os.path.exists(file_path):
        # Open the saved file for reading
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=assets.xlsx'
        return response
    else:
        return HttpResponse("File not found.")

# @login_required
# def export_to_excel(request):
#     # Create a new Excel workbook and select the active worksheet
#     wb = Workbook()
#     ws = wb.active
    
#     # Write the headers for the Excel file
#     headers = ['asset_id', 'name', 'category', 'location', 'owner', 'purchase_date']
#     ws.append(headers)
    
#     # Write the asset data to the Excel file
#     assets = Asset.objects.all().values_list(
#         'asset_id', 'name', 'category__category_name', 'location', 'owner', 'purchase_date'
#     )
#     for asset in assets:
#         ws.append(asset)
    
#     # Save the Excel file to a specific location
#     file_path = os.path.join(settings.BASE_DIR, 'assets.xlsx')
#     wb.save(file_path)
    
#     # Check if the file exists
#     if os.path.exists(file_path):
#         # Open the saved file for reading
#         with open(file_path, 'rb') as file:
#             response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
#             response['Content-Disposition'] = 'attachment; filename=assets.xlsx'
#         return response
#     else:
#         return HttpResponse("File not found.")
    
# @login_required
# def export_to_excel(request):
#     # Create a new Excel workbook and select the active worksheet
#     wb = Workbook()
#     ws = wb.active
    
#     # Write the headers for the Excel file
#     headers = ['asset_id', 'name', 'category', 'sub_category', 'location', 'owner', 'purchase_date']
#     ws.append(headers)
    
#     # Write the asset data to the Excel file
#     assets = Asset.objects.all().values_list(
#         'asset_id', 'name', 'category__category_name', 'sub_category', 'location', 'owner', 'purchase_date'
#     )
#     for asset in assets:
#         ws.append(asset)
    
#     # Save the Excel file to a specific location
#     file_path = os.path.join(settings.BASE_DIR, 'assets.xlsx')
#     wb.save(file_path)
    
#     # Check if the file exists
#     if os.path.exists(file_path):
#         # Open the saved file for reading
#         with open(file_path, 'rb') as file:
#             response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
#             response['Content-Disposition'] = 'attachment; filename=assets.xlsx'
#         return response
#     else:
#         return HttpResponse("File not found.")
    

# print qr code in bulk
@login_required
def print_qr_codes(request):
    # Get all active assets
    assets = Asset.objects.filter(is_active=True)

    # Create a list to store the data for the QR codes
    qr_code_data = []
    for asset in assets:
        qr_code_data.append({
            'asset_id': asset.asset_id,
            'qr_code_url': asset.qr_code.url,
            'name': asset.name,
            'category': asset.category,
        })

    # Generate the PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="qr_codes.pdf"'

    p = canvas.Canvas(response, pagesize=A4)

    qr_code_width = 80
    qr_code_height = 80
    x_start = 25
    y_start = 750
    x = x_start
    y = y_start
    line_spacing = 37
    qr_code_per_page = 35

    for i, data in enumerate(qr_code_data):
        if i > 0 and i % qr_code_per_page == 0:
            p.showPage()
            x = x_start
            y = y_start

        if i > 0 and i % 5 == 0:
            x = x_start
            y -= qr_code_height + line_spacing

        if i > 0 and i % 35 == 0:
            x = x_start
            y = y_start

        # Draw the QR code
        p.drawImage(data['qr_code_url'], x, y, width=qr_code_width, height=qr_code_height)

        # Draw the asset ID and location
        p.setFont("Helvetica", 6)
        p.drawString(x+5, y - 7, f"Asset ID: {data['asset_id']}")
        p.drawString(x+19, y - 16, f"Name: {data['name']}")
        p.drawString(x+5, y - 25, f"Category: {data['category']}")

        x += qr_code_width + line_spacing

    p.save()
    return response

@login_required
def print_qr_codes_with_loading(request):
    # Render a template with a loading message
    loading_context = {
        'loading_message': 'Please wait while the PDF is being generated...'
    }
    return render(request, 'print_qr_codes_loading.html', loading_context, content_type='text/html')

# Export report as excel

@login_required
def export_report_as_excel(request):
    # Fetch all report objects
    reports = Report.objects.all()

    # Create a new workbook and get the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Define the headers
    headers = ['Name', 'Email', 'Description', 'Asset ID', 'Service Type', 'Remark', 'Report Date']

    # Write the headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Write the report data to the worksheet
    for row_num, report in enumerate(reports, 2):
        worksheet.cell(row=row_num, column=1).value = report.name
        worksheet.cell(row=row_num, column=2).value = report.email
        worksheet.cell(row=row_num, column=3).value = report.description
        worksheet.cell(row=row_num, column=4).value = report.asset.asset_id
        worksheet.cell(row=row_num, column=5).value = report.service_type
        worksheet.cell(row=row_num, column=6).value = report.remark
        worksheet.cell(row=row_num, column=7).value = report.item_creation_date.replace(tzinfo=None)

    # Set the response headers for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

# inactive asset view
from django.db.models import Q
from django.shortcuts import render
from backend.models import Asset

@login_required
def inactive_assets(request):
    user = request.user
    role_name = None

    if user.is_superuser:
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        role_name = 'ICT'

    search_term = request.GET.get('search_term')
    inactive_assets = Asset.objects.filter(is_active=False).order_by('-item_creation_date')

    if search_term:
        inactive_assets = inactive_assets.filter(Q(asset_id__icontains=search_term))

    paginator = Paginator(inactive_assets, 10)  # Show 10 inactive assets per page

    page = request.GET.get('page')
    try:
        inactive_assets = paginator.page(page)
    except PageNotAnInteger:
        inactive_assets = paginator.page(1)
    except EmptyPage:
        inactive_assets = paginator.page(paginator.num_pages)

    num_pages = paginator.num_pages
    current_page = inactive_assets.number

    if num_pages <= 5:
        page_range = range(1, num_pages + 1)
    elif current_page <= 3:
        page_range = range(1, 6)
    elif current_page >= num_pages - 2:
        page_range = range(num_pages - 4, num_pages + 1)
    else:
        page_range = range(current_page - 2, current_page + 3)

    context = {
        'inactive_assets': inactive_assets,
        'page_range': page_range,
        'category': AssetCategory.objects.all(),
        'search_term': search_term,
        'role_name': role_name,
    }

    return render(request, 'inactive.html', context)

# activate asset:
@login_required
def activate_asset(request, asset_id):
    asset = get_object_or_404(Asset, asset_id=asset_id)
    
    try:
        inactive_asset = InactiveAsset.objects.get(asset=asset)
        inactive_asset.delete()  # Remove the asset from the inactive assets
    except InactiveAsset.DoesNotExist:
        pass  # No inactive asset exists for the given asset

    # Set is_active to True
    asset.is_active = True  
    asset.save()

    messages.success(request, 'Asset has been activated.')

    # Perform any additional actions or redirects as needed
    return redirect('backend:inactive_assets')


#export inactive asset as excel:
@login_required
def export_inactive_to_excel(request):
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write the headers for the Excel file
    headers = ['asset_id', 'name', 'category', 'location', 'owner', 'purchase_date']
    ws.append(headers)

    # Write the inactive asset data to the Excel file
    inactive_assets = Asset.objects.filter(is_active=False).values_list(
        'asset_id', 'name', 'category__category_name', 'location', 'owner', 'purchase_date'
    )
    for asset in inactive_assets:
        ws.append(asset)

    # Save the Excel file to a specific location
    file_path = os.path.join(settings.BASE_DIR, 'inactive_assets.xlsx')
    wb.save(file_path)

    # Check if the file exists
    if os.path.exists(file_path):
        # Open the saved file for reading
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=inactive_assets.xlsx'
        return response
    else:
        return HttpResponse("File not found.")


# move report to report log after service is done
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from backend.models import Report, ReportLog

@login_required
def move_report_to_log(request, report_id):
    report = get_object_or_404(Report, id=report_id)

    # Create a new ReportLog instance with all the report details
    report_log = ReportLog(
        name=report.name,
        email=report.email,
        description=report.description,
        asset=report.asset,
        service_type=report.service_type,
        remark=report.remark,
        item_creation_date=report.item_creation_date,
        completion_date=report.item_creation_date,
        completed_by=request.user
    )
    report_log.save()

    # Delete the original report
    report.delete()

    messages.success(request, 'Report has beeen saved to logs.')
    return redirect('backend:report')


# report log view 
@login_required
def report_log(request):
    category = AssetCategory.objects.all()
    user = request.user

    show_reports = False
    hide_action_column = False  # Flag to determine if the action column should be hidden
    role_name = None

    if user.is_superuser and user.is_staff:
        # Show all report logs for superadmin and staff users
        report_log_list = ReportLog.objects.order_by('completion_date').all()
        show_reports = True
        role_name = 'Admin'
    elif user.is_staff and not user.groups.exists():
        # Show all report logs for staff users who are not in any group
        report_log_list = ReportLog.objects.order_by('completion_date').all()
        show_reports = True
        hide_action_column = True  # Hide the action column for these users
        role_name = 'Staff'
    elif user.groups.filter(name='state_user').exists():
        # Show only estate service report logs for staff users in state_user group
        report_log_list = ReportLog.objects.filter(service_type='estate').order_by('completion_date').all()
        show_reports = True
        role_name = 'Estate'
    elif user.groups.filter(name='ict_user').exists():
        # Show only ICT service report logs for staff users in ict_user group
        report_log_list = ReportLog.objects.filter(service_type='ict').order_by('completion_date').all()
        show_reports = True
        role_name = 'ICT'
    else:
        # For other users, show no report logs
        report_log_list = ReportLog.objects.none()

    # Handle search functionality
    search_term = request.GET.get('search_term')
    if search_term:
        report_log_list = report_log_list.filter(asset__asset_id__icontains=search_term)

    paginator = Paginator(report_log_list, 10)  # Show 10 report logs per page
    page = request.GET.get('page')
    report_logs = paginator.get_page(page)

    # Calculate the range of pages to display
    num_pages = paginator.num_pages
    current_page = report_logs.number

    if num_pages <= 5:
        page_range = range(1, num_pages + 1)
    elif current_page <= 3:
        page_range = range(1, 6)
    elif current_page >= num_pages - 2:
        page_range = range(num_pages - 4, num_pages + 1)
    else:
        page_range = range(current_page - 2, current_page + 3)

    prev_page = report_logs.previous_page_number if report_logs.has_previous() else None
    next_page = report_logs.next_page_number if report_logs.has_next() else None

    context = {
        'category': category,
        'report_logs': report_logs,
        'search_term': search_term,
        'page_range': page_range,
        'prev_page': prev_page,
        'next_page': next_page,
        'show_reports': show_reports,
        'hide_action_column': hide_action_column,
        'role_name': role_name,
    }

    return render(request, 'report_log.html', context)

# def report_log(request):
#     category = AssetCategory.objects.all()
#     user = request.user
#     show_reports = False
#     hide_action_column = False  # Flag to determine if the action column should be hidden

#     if user.is_superuser and user.is_staff:
#         # Show all report logs for superadmin and staff users
#         report_log_list = ReportLog.objects.order_by('completion_date').all()
#         show_reports = True
#     elif user.is_staff and not user.groups.exists():
#         # Show all report logs for staff users who are not in any group
#         report_log_list = ReportLog.objects.order_by('completion_date').all()
#         show_reports = True
#         hide_action_column = True  # Hide the action column for these users
  
#     elif user.groups.filter(name='state_user').exists():
#     # Show only estate service report logs for staff users in state_user group
#         report_log_list = ReportLog.objects.filter(service_type='estate')
#         show_reports = True

    
#     elif user.groups.filter(name='ict_user').exists():
#     # Show only ICT service report logs for staff users in ict_user group
#         report_log_list = ReportLog.objects.filter(service_type='ict')
#         show_reports = True

#     else:
#         # For other users, show no report logs
#         report_log_list = ReportLog.objects.none()

#     # Handle search functionality
#     search_term = request.GET.get('search_term')
#     if search_term:
#         report_log_list = report_log_list.filter(asset__asset_id__icontains=search_term)

#     paginator = Paginator(report_log_list, 10)  # Show 10 report logs per page
#     page = request.GET.get('page')
#     report_logs = paginator.get_page(page)

#     # Calculate the range of pages to display
#     num_pages = paginator.num_pages
#     current_page = report_logs.number

#     if num_pages <= 5:
#         page_range = range(1, num_pages + 1)
#     elif current_page <= 3:
#         page_range = range(1, 6)
#     elif current_page >= num_pages - 2:
#         page_range = range(num_pages - 4, num_pages + 1)
#     else:
#         page_range = range(current_page - 2, current_page + 3)

#     prev_page = report_logs.previous_page_number if report_logs.has_previous() else None
#     next_page = report_logs.next_page_number if report_logs.has_next() else None

#     context = {
#         'category': category,
#         'report_logs': report_logs,
#         'search_term': search_term,
#         'page_range': page_range,
#         'prev_page': prev_page,
#         'next_page': next_page,
#         'show_reports': show_reports,
#         'hide_action_column': hide_action_column, 
#     }
#     return render(request, 'report_log.html', context)


@login_required
def delete_log(request, report_log_id):
    if request.method == 'POST':
        reportLog = get_object_or_404(ReportLog, id=report_log_id)
        reportLog.delete()

        messages.success(request, 'Report Log deleted successfully')
        return redirect('backend:report_log')


#export report log to excel
@login_required
def export_report_log_as_excel(request):
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define cell styles
    date_format = NamedStyle(name='date_format')
    date_format.number_format = 'dd/mm/yyyy'
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Write the headers for the Excel file
    headers = ['Asset ID', 'Name', 'Email', 'Description', 'Report Date', 'Remark', 'Completed By', 'Date of Completion']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[col_letter + '1'] = header
        ws[col_letter + '1'].font = bold_font
        ws[col_letter + '1'].fill = header_fill
        ws.column_dimensions[col_letter].width = 15

    # Write the report log data to the Excel file
    report_logs = ReportLog.objects.all().values_list(
        'asset__asset_id', 'name', 'email', 'description', 'item_creation_date', 'remark', 'completed_by', 'completion_date'
    )
    for row_num, log in enumerate(report_logs, 2):
        for col_num, field in enumerate(log, 1):
            col_letter = get_column_letter(col_num)
            # Apply cell style to date fields
            if isinstance(field, datetime):
                field = field.astimezone(pytz.timezone('UTC')).replace(tzinfo=None)
                ws[col_letter + str(row_num)].style = date_format
            # Convert completed_by user ID to actual name
            if col_num == 7:  # Column index of 'Completed By'
                completed_by_id = field
                completed_by_name = User.objects.get(id=completed_by_id).username
                ws[col_letter + str(row_num)] = completed_by_name
            else:
                ws[col_letter + str(row_num)] = field

    # Save the Excel file to a specific location
    file_path = os.path.join(settings.BASE_DIR, 'report_logs.xlsx')
    wb.save(file_path)

    # Check if the file exists
    if os.path.exists(file_path):
        # Open the saved file for reading
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=report_logs.xlsx'
        return response
    else:
        return HttpResponse("File not found.")

























    




